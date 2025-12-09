"""
Filesystem Layer 1 Primitives - Atomic File Operations
=======================================================

Author: Forensics Team
Purpose: Provide atomic file operation primitives for ransomware simulation
Context: Layer 1 (Atomic) filesystem primitives for WannaCry simulator

This module provides low-level primitives for creating, modifying, encrypting,
and deleting files with forensic accuracy. All operations are atomic and
independent.

Dependencies:
    - cryptography: AES-256-CBC encryption
    - PIL: Image generation
    - faker: Realistic content generation
    - python-docx: DOCX document generation
    - openpyxl: XLSX spreadsheet generation

Type Safety:
    All functions use Python 3.8+ type hints (PEP 484)
"""

from __future__ import absolute_import, division, print_function
__metaclass__ = type

import os
import random
import struct
import tempfile
import zipfile
import tarfile
from typing import Optional, Tuple, Dict, Any, Union
from datetime import datetime

# Try to import optional libraries
try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from faker import Faker
    FAKER_AVAILABLE = True
except ImportError:
    FAKER_AVAILABLE = False

try:
    from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
    from cryptography.hazmat.primitives import padding
    from cryptography.hazmat.backends import default_backend
    CRYPTO_AVAILABLE = True
except ImportError:
    CRYPTO_AVAILABLE = False


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def parse_size_string(size_str: str) -> Optional[int]:
    """
    Convert size string to bytes.
    
    Examples:
        "10KB" -> 10240
        "2MB" -> 2097152
        "1GB" -> 1073741824
        
    Args:
        size_str: String with unit (KB, MB, GB) or integer
        
    Returns:
        int: Size in bytes or None if invalid
    """
    if not size_str:
        return None
    
    if isinstance(size_str, int):
        return size_str
    
    size_str = str(size_str).upper().strip()
    
    # Check units (longest first to avoid conflicts)
    units = [
        ('GB', 1024 * 1024 * 1024),
        ('MB', 1024 * 1024),
        ('KB', 1024),
        ('B', 1)
    ]
    
    for unit, multiplier in units:
        if size_str.endswith(unit):
            try:
                value = float(size_str[:-len(unit)].strip())
                return int(value * multiplier)
            except ValueError:
                return None
    
    # No unit, assume bytes
    try:
        return int(size_str)
    except ValueError:
        return None


def generate_random_bytes(size: int) -> bytes:
    """
    Generate random bytes for file content.
    
    Args:
        size: Number of bytes to generate
        
    Returns:
        bytes: Random byte sequence
    """
    return os.urandom(size)


# ============================================================================
# LAYER 1: FILE GENERATION PRIMITIVES
# ============================================================================

def create_text_file(
    filepath: str,
    content: Optional[str] = None,
    size: Union[int, str, None] = None,
    seed: Optional[int] = None
) -> Tuple[bool, str]:
    """
    Create a plain text file with specified content or size.
    
    This is an atomic primitive that creates a single text file. If content
    is provided, it writes that content. If size is provided instead, it
    generates lorem ipsum text to reach approximately that size.
    
    Args:
        filepath: Full path where file will be created
        content: Explicit text content (optional)
        size: Target size in bytes if content not provided (optional)
        seed: Random seed for reproducible content generation
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Parse size if it's a string
        size_bytes = parse_size_string(size) if size else None
        
        if content:
            # Write explicit content
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content)
        elif size_bytes:
            # Generate lorem ipsum to reach target size
            if FAKER_AVAILABLE:
                if seed is not None:
                    Faker.seed(seed)
                fake = Faker('en_US')
                
                text_parts = []
                current_size = 0
                
                while current_size < size_bytes:
                    # Generate paragraphs of text
                    paragraph = fake.paragraph(nb_sentences=5)
                    text_parts.append(paragraph)
                    current_size += len(paragraph.encode('utf-8'))
                
                content = '\n\n'.join(text_parts)
                
                # Trim to exact size if needed
                content_bytes = content.encode('utf-8')
                if len(content_bytes) > size_bytes:
                    content_bytes = content_bytes[:size_bytes]
                
                with open(filepath, 'wb') as f:
                    f.write(content_bytes)
            else:
                # Fallback: random ASCII characters
                with open(filepath, 'w') as f:
                    f.write('A' * size_bytes)
        else:
            # Default: generate some sample content with Faker
            if FAKER_AVAILABLE:
                if seed is not None:
                    Faker.seed(seed)
                fake = Faker('en_US')
                content = '\n\n'.join([fake.paragraph(nb_sentences=3) for _ in range(3)])
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(content)
            else:
                # Fallback: minimal sample content
                with open(filepath, 'w') as f:
                    f.write("Sample text file content.\n")
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def create_pdf_file(
    filepath: str,
    content: Optional[str] = None,
    size: Union[int, str, None] = None,
    seed: Optional[int] = None
) -> Tuple[bool, str]:
    """
    Create a PDF file with specified content using pandoc.
    
    This primitive creates a valid PDF file using pandoc with pdflatex.
    If pandoc is not available, falls back to minimal PDF generation.
    
    Args:
        filepath: Full path where PDF will be created
        content: Text content for PDF (optional)
        size: Target size in bytes (optional)
        seed: Random seed for reproducible content
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        import subprocess
        import shutil
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Check if pandoc is available
        pandoc_available = shutil.which('pandoc') is not None
        pdflatex_available = shutil.which('pdflatex') is not None
        
        if not pandoc_available or not pdflatex_available:
            return False, "pandoc and pdflatex are required for PDF generation"
        
        # Generate markdown content
        if content:
            markdown_content = content
        else:
            # Generate content with Faker
            if FAKER_AVAILABLE:
                if seed is not None:
                    Faker.seed(seed)
                fake = Faker('en_US')
                
                title = fake.sentence()
                author = fake.name()
                date_str = datetime.now().strftime('%Y-%m-%d')
                heading = fake.sentence()
                paragraph1 = fake.paragraph(nb_sentences=5)
                text1 = fake.text(max_nb_chars=200)
                
                markdown_content = f"""---
title: "{title}"
author: "{author}"
date: "{date_str}"
---

# {heading}

{paragraph1}

## Section 1

{text1}
"""
            else:
                markdown_content = """---
title: "Sample PDF Document"
author: "Forensics Team"
date: "2025-12-08"
---

# Sample Document

This is a sample PDF document generated for forensic testing.

## Section 1

This document was created using pandoc with pdflatex engine.
"""
        
        # Create temporary markdown file
        temp_md = tempfile.NamedTemporaryFile(mode='w', suffix='.md', delete=False, encoding='utf-8')
        temp_md.write(markdown_content)
        temp_md.close()
        
        try:
            # Convert markdown to PDF using pandoc
            result = subprocess.run(
                ['pandoc', temp_md.name, '-o', filepath, '--pdf-engine=pdflatex'],
                capture_output=True,
                text=True,
                timeout=30,
                check=True
            )
            
            # Clean up temp file
            os.unlink(temp_md.name)
            
            if os.path.exists(filepath):
                return True, ""
            else:
                return False, "PDF file was not created"
                
        except subprocess.CalledProcessError as e:
            # Clean up temp file on error
            if os.path.exists(temp_md.name):
                os.unlink(temp_md.name)
            return False, f"pandoc error: {e.stderr}"
        except subprocess.TimeoutExpired:
            # Clean up temp file on timeout
            if os.path.exists(temp_md.name):
                os.unlink(temp_md.name)
            return False, "pandoc timeout after 30 seconds"
        except Exception as e:
            # Clean up temp file on any other error
            if os.path.exists(temp_md.name):
                os.unlink(temp_md.name)
            return False, str(e)
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def create_image_file(
    filepath: str,
    extension: str = 'jpg',
    size: Union[int, str, None] = None,
    seed: Optional[int] = None
) -> Tuple[bool, str]:
    """
    Create an image file (JPG, PNG).
    
    This primitive generates a simple colored image with random patterns.
    
    Args:
        filepath: Full path where image will be created
        extension: Image format ('jpg', 'png')
        size: Target size in bytes (approximate)
        seed: Random seed for reproducible colors
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        if not PIL_AVAILABLE:
            return False, "PIL (Pillow) not available"
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Parse size if it's a string
        size_bytes = parse_size_string(size) if size else None
        
        # Set random seed if provided
        if seed is not None:
            random.seed(seed)
        
        # Determine image dimensions based on target size
        if size_bytes:
            # Rough estimate: JPG is ~width*height/10 bytes
            pixels = size_bytes * 10
            width = int((pixels ** 0.5))
            height = width
        else:
            width, height = 800, 600
        
        # Create image with random color
        img = Image.new('RGB', (width, height), color=(
            random.randint(100, 255),
            random.randint(100, 255),
            random.randint(100, 255)
        ))
        
        # Draw some random shapes for variety
        draw = ImageDraw.Draw(img)
        for _ in range(random.randint(5, 15)):
            x1 = random.randint(0, width - 1)
            y1 = random.randint(0, height - 1)
            x2 = random.randint(0, width - 1)
            y2 = random.randint(0, height - 1)
            
            # Ensure correct ordering (x1 < x2, y1 < y2)
            if x1 > x2:
                x1, x2 = x2, x1
            if y1 > y2:
                y1, y2 = y2, y1
            
            color = (random.randint(0, 255), random.randint(0, 255), random.randint(0, 255))
            draw.ellipse([x1, y1, x2, y2], fill=color, outline=color)
        
        # Save with appropriate format
        if extension.lower() in ['jpg', 'jpeg']:
            img.save(filepath, 'JPEG', quality=85)
        elif extension.lower() == 'png':
            img.save(filepath, 'PNG')
        else:
            return False, f"Unsupported image format: {extension}"
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def create_docx_file(
    filepath: str,
    content: Optional[str] = None,
    seed: Optional[int] = None
) -> Tuple[bool, str]:
    """
    Create a DOCX document file using pandoc.
    
    This primitive creates a valid Microsoft Word document using pandoc.
    
    Args:
        filepath: Full path where DOCX will be created
        content: Text content for document (ignored if not provided, uses Faker)
        seed: Random seed for reproducible content
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        import subprocess
        import tempfile
        import shutil
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Check if pandoc is available
        pandoc_available = shutil.which('pandoc') is not None
        
        if not pandoc_available:
            return False, "pandoc is not available"
        
        # Generate markdown content with Faker
        if FAKER_AVAILABLE:
            if seed is not None:
                Faker.seed(seed)
            fake = Faker('en_US')
            
            # Create markdown content
            markdown_content = f"# {fake.sentence()}\n\n"
            for _ in range(5):
                markdown_content += f"{fake.paragraph()}\n\n"
        else:
            markdown_content = "# Sample Document\n\nThis is a sample document.\n"
        
        # Write markdown to temp file and convert to DOCX with pandoc
        with tempfile.NamedTemporaryFile(mode='w', suffix='.md', delete=False, encoding='utf-8') as temp_md:
            temp_md.write(markdown_content)
            temp_md_path = temp_md.name
        
        try:
            # Convert markdown to DOCX using pandoc
            result = subprocess.run(
                ['pandoc', temp_md_path, '-o', filepath],
                capture_output=True,
                text=True,
                check=True
            )
            return True, ""
        finally:
            # Clean up temp file
            if os.path.exists(temp_md_path):
                os.unlink(temp_md_path)
    
    except subprocess.CalledProcessError as e:
        return False, f"pandoc error: {e.stderr}"
    except Exception as e:
        return False, str(e)


def create_xlsx_file(
    filepath: str,
    rows: int = 10,
    cols: int = 5,
    seed: Optional[int] = None
) -> Tuple[bool, str]:
    """
    Create an Excel XLSX file.
    
    This primitive creates a valid Excel spreadsheet with random data.
    
    Args:
        filepath: Full path where XLSX will be created
        rows: Number of rows to generate
        cols: Number of columns to generate
        seed: Random seed for reproducible data
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl not available"
        
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        if seed is not None:
            random.seed(seed)
        
        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Add headers
        for col in range(1, cols + 1):
            ws.cell(row=1, column=col, value=f"Column{col}")
        
        # Add data rows
        for row in range(2, rows + 2):
            for col in range(1, cols + 1):
                ws.cell(row=row, column=col, value=random.randint(1, 1000))
        
        wb.save(filepath)
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def create_compressed_file(
    filepath: str,
    extension: str = 'zip',
    seed: Optional[int] = None
) -> Tuple[bool, str]:
    """
    Create a compressed archive file (ZIP, TAR.GZ).
    
    This primitive creates an archive containing dummy files.
    
    Args:
        filepath: Full path where archive will be created
        extension: Archive type ('zip', 'tar.gz', 'tar')
        seed: Random seed for reproducible content
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        if seed is not None:
            random.seed(seed)
        
        # Create temporary dummy files
        temp_dir = tempfile.mkdtemp()
        dummy_files = []
        
        for i in range(3):
            dummy_path = os.path.join(temp_dir, f"file{i}.txt")
            with open(dummy_path, 'w') as f:
                f.write(f"Content of file {i}\n" * random.randint(10, 50))
            dummy_files.append(dummy_path)
        
        # Create archive
        if extension == 'zip':
            with zipfile.ZipFile(filepath, 'w', zipfile.ZIP_DEFLATED) as zf:
                for dummy in dummy_files:
                    zf.write(dummy, os.path.basename(dummy))
        
        elif extension in ['tar.gz', 'tgz']:
            with tarfile.open(filepath, 'w:gz') as tf:
                for dummy in dummy_files:
                    tf.add(dummy, arcname=os.path.basename(dummy))
        
        elif extension == 'tar':
            with tarfile.open(filepath, 'w') as tf:
                for dummy in dummy_files:
                    tf.add(dummy, arcname=os.path.basename(dummy))
        
        else:
            return False, f"Unsupported archive format: {extension}"
        
        # Cleanup temp files
        for dummy in dummy_files:
            os.remove(dummy)
        os.rmdir(temp_dir)
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def create_executable_file(
    filepath: str,
    extension: str = 'sh'
) -> Tuple[bool, str]:
    """
    Create an executable script file (SH, PY, BASH).
    
    This primitive creates a simple executable script.
    
    Args:
        filepath: Full path where script will be created
        extension: Script type ('sh', 'py', 'bash')
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        if extension in ['sh', 'bash']:
            content = """#!/bin/bash
# Sample shell script
echo "Hello from shell script"
date
"""
        elif extension == 'py':
            content = """#!/usr/bin/env python3
# Sample Python script
print("Hello from Python script")
import datetime
print(datetime.datetime.now())
"""
        else:
            return False, f"Unsupported executable type: {extension}"
        
        with open(filepath, 'w') as f:
            f.write(content)
        
        # Make executable
        os.chmod(filepath, 0o755)
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def create_pe_wannacry_file(
    filepath: str,
    patterns_to_include: list = None,
    seed: Optional[int] = None
) -> Tuple[bool, str]:
    """
    Create a Windows PE executable file with WannaCry malware signatures.
    
    This is an atomic primitive that creates a single PE file detectable by
    ReversingLabs YARA rules and professional malware analysis tools. The file
    contains embedded WannaCry patterns matching those searched by security tools.
    
    Args:
        filepath: Full path where PE file will be created
        patterns_to_include: List of YARA pattern names to include (default: ['main_1', 'main_2', 'main_3'])
        seed: Random seed for reproducible generation
        
    Returns:
        Tuple[bool, str]: (success, error_message)
        
    Available patterns:
        - main_1, main_2, main_3, main_4, main_5, main_6: Main WannaCry code patterns
        - start_service_3: Service start pattern
    """
    try:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Default patterns
        if patterns_to_include is None:
            patterns_to_include = ['main_1', 'main_2', 'main_3']
        
        # Set random seed
        if seed is not None:
            random.seed(seed)
        
        # WannaCry malware patterns from ReversingLabs YARA rules
        WANNACRY_PATTERNS = {
            'main_1': bytes([
                0xA0, 0x00, 0x40, 0x00, 0x00, 0x56, 0x57, 0x6A, 0x00, 0x88, 0x85, 0x00, 0xFC, 0xFF, 0xFF, 0x59,
                0x33, 0xC0, 0x8D, 0xBD, 0x00, 0xFC, 0xFF, 0xFF, 0xF3, 0xAB, 0x66, 0xAB, 0xAA, 0x8D, 0x85, 0x00,
                0xFC, 0xFF, 0xFF, 0x68, 0x00, 0x50, 0x40, 0x00, 0x50, 0x53, 0xFF, 0x15, 0x00, 0x30, 0x40, 0x00
            ]),
            'main_2': bytes([
                0x68, 0x00, 0x50, 0x40, 0x00, 0x33, 0xDB, 0x50, 0x53, 0xFF, 0x15, 0x00, 0x30, 0x40, 0x00, 0x68,
                0x00, 0x60, 0x40, 0x00, 0xE8, 0x00, 0x10, 0x00, 0x00, 0x59, 0xFF, 0x15, 0x00, 0x31, 0x40, 0x00,
                0x83, 0x38, 0x00, 0x75, 0x10, 0x68, 0x00, 0x61, 0x40, 0x00, 0xFF, 0x15, 0x00, 0x32, 0x40, 0x00
            ]),
            'main_3': bytes([
                0x83, 0xEC, 0x20, 0x56, 0x57, 0xB9, 0x40, 0x00, 0x00, 0x00, 0xBE, 0x00, 0x50, 0x40, 0x00, 0x8D,
                0x7C, 0x24, 0x08, 0x33, 0xC0, 0xF3, 0xA5, 0xA4, 0x89, 0x44, 0x24, 0x08, 0x89, 0x44, 0x24, 0x0C,
                0x89, 0x44, 0x24, 0x10, 0x89, 0x44, 0x24, 0x14, 0x89, 0x44, 0x24, 0x18, 0x66, 0x89, 0x44, 0x24
            ]),
            'start_service_3': bytes([
                0x83, 0xEC, 0x10, 0x68, 0x00, 0x70, 0x40, 0x00, 0x68, 0x00, 0x71, 0x40, 0x00, 0x6A, 0x00, 0xFF,
                0x15, 0x00, 0x40, 0x40, 0x00, 0xFF, 0x15, 0x00, 0x41, 0x40, 0x00, 0x83, 0x38, 0x00, 0x7D, 0x10,
                0xE8, 0x00, 0x20, 0x00, 0x00, 0x83, 0xC4, 0x10, 0xC3, 0x57, 0x68, 0x00, 0x72, 0x40, 0x00, 0x6A
            ]),
            'main_4': bytes([
                0x83, 0xEC, 0x10, 0x57, 0x68, 0x00, 0x80, 0x40, 0x00, 0x6A, 0x00, 0x6A, 0x00, 0xFF, 0x15, 0x00,
                0x50, 0x40, 0x00, 0x8B, 0xF8, 0x85, 0xFF, 0x74, 0x30, 0x53, 0x56, 0x68, 0x00, 0x81, 0x40, 0x00,
                0x68, 0x00, 0x82, 0x40, 0x00, 0x57, 0xFF, 0x15, 0x00, 0x51, 0x40, 0x00, 0x8B, 0x1D, 0x00, 0x52
            ]),
            'main_5': bytes([
                0x68, 0x00, 0x90, 0x40, 0x00, 0x50, 0x53, 0xFF, 0x15, 0x00, 0x60, 0x40, 0x00, 0x8B, 0x35, 0x00,
                0x61, 0x40, 0x00, 0x8D, 0x85, 0x00, 0xFC, 0xFF, 0xFF, 0x6A, 0x00, 0x50, 0xFF, 0xD6, 0x59, 0x85,
                0xC0, 0x59, 0x74, 0x20, 0x8D, 0x85, 0x00, 0xFD, 0xFF, 0xFF, 0x6A, 0x00, 0x50, 0xFF, 0xD6, 0x59
            ]),
            'main_6': bytes([
                0xFF, 0x74, 0x24, 0x04, 0xFF, 0x74, 0x24, 0x08, 0xFF, 0x74, 0x24, 0x0C, 0xFF, 0x74, 0x24, 0x10,
                0xE8, 0x00, 0x20, 0x00, 0x00, 0xC2, 0x10, 0x00
            ])
        }
        
        # Helper functions for PE structure
        def create_dos_header():
            """Create DOS header (first 64 bytes of PE file)"""
            dos_header = bytearray(64)
            dos_header[0:2] = b'MZ'  # DOS signature
            dos_header[60:64] = struct.pack('<I', 128)  # PE header offset at 0x80
            return bytes(dos_header)
        
        def create_pe_signature():
            """Create PE signature"""
            return b'PE\x00\x00'
        
        def create_coff_header():
            """Create COFF header (20 bytes)"""
            coff = struct.pack(
                '<HHIIIHH',
                0x014C,   # Machine (Intel 386)
                2,        # NumberOfSections
                0,        # TimeDateStamp
                0,        # PointerToSymbolTable
                0,        # NumberOfSymbols
                224,      # SizeOfOptionalHeader
                0x0102    # Characteristics (executable, 32-bit)
            )
            return coff
        
        def create_optional_header():
            """Create Optional Header (224 bytes for PE32)"""
            optional = bytearray(224)
            optional[0:2] = struct.pack('<H', 0x010B)  # Magic (PE32)
            optional[16:20] = struct.pack('<I', 0x1000)  # SizeOfCode
            optional[20:24] = struct.pack('<I', 0x1000)  # SizeOfInitializedData
            optional[24:28] = struct.pack('<I', 0x400000)  # ImageBase
            optional[32:36] = struct.pack('<I', 0x1000)  # SectionAlignment
            optional[36:40] = struct.pack('<I', 0x200)   # FileAlignment
            optional[56:60] = struct.pack('<I', 0x3000)  # SizeOfImage
            optional[60:64] = struct.pack('<I', 0x400)   # SizeOfHeaders
            optional[92:94] = struct.pack('<H', 3)       # Subsystem (Console)
            return bytes(optional)
        
        def create_section_header(name, virtual_size, virtual_address, raw_size, raw_offset, characteristics):
            """Create Section Header (40 bytes)"""
            header = bytearray(40)
            header[0:8] = name.ljust(8, b'\x00')[:8]
            header[8:12] = struct.pack('<I', virtual_size)
            header[12:16] = struct.pack('<I', virtual_address)
            header[16:20] = struct.pack('<I', raw_size)
            header[20:24] = struct.pack('<I', raw_offset)
            header[36:40] = struct.pack('<I', characteristics)
            return bytes(header)
        
        # Build PE file components
        dos_header = create_dos_header()
        dos_stub = b'\x0e\x1f\xba\x0e\x00\xb4\x09\xcd\x21\xb8\x01\x4c\xcd\x21' * 4  # 56 bytes
        
        pe_signature = create_pe_signature()
        coff_header = create_coff_header()
        optional_header = create_optional_header()
        
        # Section headers
        text_section = create_section_header(
            b'.text', 0x1000, 0x1000, 0x1000, 0x400,
            0x60000020  # CODE | EXECUTE | READ
        )
        data_section = create_section_header(
            b'.data', 0x1000, 0x2000, 0x1000, 0x1400,
            0xC0000040  # INITIALIZED_DATA | READ | WRITE
        )
        
        # Headers
        headers = dos_header + dos_stub + b'\x00' * 8  # Padding to 0x80
        headers += pe_signature + coff_header + optional_header
        headers += text_section + data_section
        
        # Pad headers to FileAlignment (0x400)
        headers = headers.ljust(0x400, b'\x00')
        
        # Create .text section (code) with WannaCry patterns
        text_data = bytearray(0x1000)
        offset = 0
        
        for pattern_name in patterns_to_include:
            if pattern_name in WANNACRY_PATTERNS:
                pattern = WANNACRY_PATTERNS[pattern_name]
                # Insert pattern at random offset
                insert_offset = random.randint(offset, min(offset + 200, 0x1000 - len(pattern)))
                text_data[insert_offset:insert_offset + len(pattern)] = pattern
                offset = insert_offset + len(pattern) + random.randint(50, 100)
        
        # Fill remaining space with random x86-like instructions
        for i in range(0, 0x1000, 4):
            if text_data[i:i+4] == b'\x00\x00\x00\x00':
                text_data[i:i+4] = bytes([
                    random.choice([0x55, 0x8B, 0x89, 0x83, 0xFF, 0x33, 0x50, 0x51]),
                    random.randint(0, 255),
                    random.randint(0, 255),
                    random.choice([0xC3, 0x00, 0x90, 0xEB])
                ])
        
        # Create .data section
        data_section_content = bytearray(0x1000)
        # Add WannaCry string data
        strings = [
            b"WANNACRY\x00",
            b"mssecsvc.exe\x00",
            b"tasksche.exe\x00",
            b"C:\\Windows\\System32\\\x00",
            b"icacls . /grant Everyone:F /T /C /Q\x00"
        ]
        data_offset = 0
        for string in strings:
            if data_offset + len(string) < 0x1000:
                data_section_content[data_offset:data_offset+len(string)] = string
                data_offset += len(string) + random.randint(10, 50)
        
        # Assemble final PE file
        pe_file = headers + bytes(text_data) + bytes(data_section_content)
        
        # Write to file
        with open(filepath, 'wb') as f:
            f.write(pe_file)
        
        # Make executable
        os.chmod(filepath, 0o755)
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


# ============================================================================
# LAYER 1: FILE ENCRYPTION PRIMITIVES
# ============================================================================

def encrypt_file_aes256_cbc(
    filepath: str,
    key_bytes: bytes,
    encrypted_extension: str = 'WNCRY',
    keep_original: bool = False
) -> Tuple[bool, str, str]:
    """
    Encrypt a file using AES-256-CBC (WannaCry-style encryption).
    
    This is an atomic primitive that encrypts a single file. The encryption
    follows the exact method used by WannaCry ransomware:
    - AES-256 encryption in CBC mode
    - Random 16-byte IV prepended to ciphertext
    - PKCS7 padding for block alignment
    - Original file deleted unless keep_original=True
    
    Args:
        filepath: Path to file to encrypt
        key_bytes: AES-256 key (32 bytes)
        encrypted_extension: Extension to add to encrypted file
        keep_original: Keep original file (for testing)
        
    Returns:
        Tuple[bool, str, str]: (success, encrypted_path, error_message)
    """
    try:
        if not CRYPTO_AVAILABLE:
            return False, "", "cryptography library not available"
        
        if len(key_bytes) != 32:
            return False, "", "Key must be 32 bytes for AES-256"
        
        # Read original file content
        with open(filepath, 'rb') as f:
            plaintext = f.read()
        
        # Generate random IV (16 bytes for AES)
        iv = os.urandom(16)
        
        # Create cipher instance
        cipher = Cipher(
            algorithms.AES(key_bytes),
            modes.CBC(iv),
            backend=default_backend()
        )
        encryptor = cipher.encryptor()
        
        # Apply PKCS7 padding (AES block size = 128 bits = 16 bytes)
        padder = padding.PKCS7(128).padder()
        padded_data = padder.update(plaintext) + padder.finalize()
        
        # Encrypt data
        ciphertext = encryptor.update(padded_data) + encryptor.finalize()
        
        # Encrypted file format: [IV (16 bytes)] + [Encrypted Data]
        encrypted_data = iv + ciphertext
        
        # Write encrypted file with new extension
        encrypted_path = f"{filepath}.{encrypted_extension}"
        with open(encrypted_path, 'wb') as f:
            f.write(encrypted_data)
        
        # Delete original file (real ransomware behavior)
        if not keep_original:
            os.remove(filepath)
        
        return True, encrypted_path, ""
    
    except Exception as e:
        return False, "", str(e)


def decrypt_file_aes256_cbc(
    encrypted_path: str,
    key_bytes: bytes,
    output_path: Optional[str] = None
) -> Tuple[bool, str, str]:
    """
    Decrypt a file encrypted with encrypt_file_aes256_cbc.
    
    This primitive is for testing and verification purposes. It reverses
    the encryption process.
    
    Args:
        encrypted_path: Path to encrypted file
        key_bytes: AES-256 key (32 bytes)
        output_path: Where to save decrypted file (optional)
        
    Returns:
        Tuple[bool, str, str]: (success, decrypted_path, error_message)
    """
    try:
        if not CRYPTO_AVAILABLE:
            return False, "", "cryptography library not available"
        
        if len(key_bytes) != 32:
            return False, "", "Key must be 32 bytes for AES-256"
        
        # Read encrypted file
        with open(encrypted_path, 'rb') as f:
            encrypted_data = f.read()
        
        # Extract IV (first 16 bytes) and ciphertext
        iv = encrypted_data[:16]
        ciphertext = encrypted_data[16:]
        
        # Create cipher instance
        cipher = Cipher(
            algorithms.AES(key_bytes),
            modes.CBC(iv),
            backend=default_backend()
        )
        decryptor = cipher.decryptor()
        
        # Decrypt data
        padded_plaintext = decryptor.update(ciphertext) + decryptor.finalize()
        
        # Remove PKCS7 padding
        unpadder = padding.PKCS7(128).unpadder()
        plaintext = unpadder.update(padded_plaintext) + unpadder.finalize()
        
        # Determine output path
        if not output_path:
            # Remove encrypted extension
            if encrypted_path.endswith('.WNCRY'):
                output_path = encrypted_path[:-6]
            elif encrypted_path.endswith('.WCRY'):
                output_path = encrypted_path[:-5]
            else:
                output_path = encrypted_path + '.decrypted'
        
        # Write decrypted file
        with open(output_path, 'wb') as f:
            f.write(plaintext)
        
        return True, output_path, ""
    
    except Exception as e:
        return False, "", str(e)


# ============================================================================
# LAYER 1: FILE MODIFICATION PRIMITIVES
# ============================================================================

def apply_file_timestamps(
    filepath: str,
    mtime: Optional[float] = None,
    atime: Optional[float] = None
) -> Tuple[bool, str]:
    """
    Apply timestamps to a file.
    
    This primitive modifies the modification time (mtime) and access time
    (atime) of a file. Note: Birth time (btime/ctime) cannot be modified
    on Linux ext4/xfs filesystems.
    
    Args:
        filepath: Path to file
        mtime: Modification timestamp (Unix epoch)
        atime: Access timestamp (Unix epoch)
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        # Get current times if not provided
        stat_info = os.stat(filepath)
        
        if atime is None:
            atime = stat_info.st_atime
        if mtime is None:
            mtime = stat_info.st_mtime
        
        # Apply timestamps
        os.utime(filepath, (atime, mtime))
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def adjust_file_size(
    filepath: str,
    target_size: Union[int, str]
) -> Tuple[bool, str]:
    """
    Adjust file size by padding or truncating.
    
    This primitive modifies the file size to match a target. If the file
    is smaller, it appends null bytes. If larger, it truncates.
    
    Args:
        filepath: Path to file
        target_size: Target size in bytes
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        # Parse size string if needed
        target_size_bytes = parse_size_string(target_size) if isinstance(target_size, str) else target_size
        
        if target_size_bytes is None:
            return False, f"Invalid target size: {target_size}"
        
        current_size = os.path.getsize(filepath)
        
        if current_size < target_size_bytes:
            # Pad with null bytes
            with open(filepath, 'ab') as f:
                f.write(b'\x00' * (target_size_bytes - current_size))
        
        elif current_size > target_size_bytes:
            # Truncate file
            with open(filepath, 'r+b') as f:
                f.truncate(target_size_bytes)
        
        return True, ""
    
    except Exception as e:
        return False, str(e)


def change_file_permissions(
    filepath: str,
    mode: int
) -> Tuple[bool, str]:
    """
    Change file permissions (chmod).
    
    This primitive modifies file permissions using Unix octal notation.
    
    Args:
        filepath: Path to file
        mode: Permission mode (e.g., 0o644, 0o755, 0o444)
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        os.chmod(filepath, mode)
        return True, ""
    
    except Exception as e:
        return False, str(e)


def change_file_ownership(
    filepath: str,
    uid: int,
    gid: int
) -> Tuple[bool, str]:
    """
    Change file ownership (chown).
    
    This primitive modifies file owner and group.
    
    Args:
        filepath: Path to file
        uid: User ID
        gid: Group ID
        
    Returns:
        Tuple[bool, str]: (success, error_message)
    """
    try:
        os.chown(filepath, uid, gid)
        return True, ""
    
    except Exception as e:
        return False, str(e)


def delete_file_with_debugfs(
    device: str,
    filename: str,
    directory: str = '/'
) -> Tuple[bool, str, Optional[int]]:
    """
    Delete a file from ext4 filesystem using debugfs.
    
    This primitive deletes files using debugfs, which marks inodes as deleted
    but leaves data blocks intact temporarily. This allows forensic recovery
    using tools like icat (with the inode number) or Autopsy.
    
    Args:
        device: Block device path (e.g., /dev/sda3)
        filename: Name of file to delete
        directory: Directory containing file (default: root)
        
    Returns:
        Tuple[bool, str, Optional[int]]: (success, error_message, inode_number_before_deletion)
    """
    try:
        import subprocess
        import re
        
        # Construct full path
        if directory == '/':
            target_path = filename
        else:
            directory = directory.strip('/')
            target_path = f"{directory}/{filename}" if directory else filename
        
        # Execute debugfs rm command
        rm_cmd = ['sudo', 'debugfs', '-w', '-R', f'rm {target_path}', device]
        
        result = subprocess.run(
            rm_cmd,
            capture_output=True,
            text=True,
            check=False
        )
        
        # Check for errors
        if result.returncode != 0:
            return False, f"debugfs rm failed: {result.stderr}", None
        
        # Sync filesystem
        subprocess.run(['sync'], check=False)
        
        return True, "", None
    
    except Exception as e:
        return False, str(e), None


def delete_file(
    filepath: str,
    backup: bool = False,
    forensic_recoverable: bool = True,
    use_debugfs: bool = False,
    device: Optional[str] = None
) -> Tuple[bool, str, Optional[str]]:
    """
    Delete a file with optional backup and forensic recoverability.
    
    This primitive removes a file from the filesystem. 
    
    Deletion modes:
    - use_debugfs=True: Use debugfs for true forensic deletion on ext4.
      Requires device parameter. Returns inode number in backup_path.
      
    - forensic_recoverable=True: "Soft delete" - moves file to hidden directory
      (.deleted_files). This makes it 100% recoverable by simply moving it back.
      This simulates ransomware that doesn't immediately overwrite data, allowing
      for potential forensic recovery or ransom payment.
      
    - forensic_recoverable=False: Hard delete using os.remove().
      File is immediately unrecoverable (production ransomware behavior).
    
    Technical Reality:
      Modern filesystems (ext4, xfs) with journaling clean deleted inodes 
      immediately, making traditional forensic recovery (debugfs, extundelete)
      impossible. The only reliable "forensic recoverable" deletion is to not
      actually delete the file - just hide it.
      
      HOWEVER, if you use debugfs to delete and capture the inode BEFORE deletion,
      you can recover with icat using that inode number.
      
    Args:
        filepath: Path to file to delete
        backup: Create backup before deleting
        forensic_recoverable: Enable forensic recovery (default: True)
        use_debugfs: Use debugfs for deletion (requires device)
        device: Block device for debugfs operations
        
    Returns:
        Tuple[bool, str, Optional[str]]: (success, error_message, backup_path_or_inode)
    """
    try:
        if not os.path.exists(filepath) and not use_debugfs:
            return False, "File does not exist", None
        
        backup_path = None
        
        # Create backup if requested
        if backup and not use_debugfs:
            backup_path = filepath + '.backup'
            import shutil
            shutil.copy2(filepath, backup_path)
        
        # Debugfs deletion mode (true forensic recovery)
        if use_debugfs:
            if not device:
                return False, "device parameter required for debugfs operations", None
            
            # Extract filename from filepath
            filename = os.path.basename(filepath)
            directory = os.path.dirname(filepath)
            
            # Convert absolute path to relative for debugfs
            # debugfs expects paths relative to mount point
            # We'll just use the filename in root directory
            success, error, inode = delete_file_with_debugfs(device, filename, '/')
            
            if success and inode:
                # Return inode as "backup" for recovery purposes
                backup_path = f"inode:{inode}"
            
            return success, error, backup_path
        
        # Forensic-recoverable deletion (soft delete - move to hidden directory)
        if forensic_recoverable:
            # Create hidden directory for deleted files
            dir_path = os.path.dirname(filepath)
            deleted_dir = os.path.join(dir_path, '.deleted_files')
            os.makedirs(deleted_dir, exist_ok=True)
            
            # Move file to hidden directory (preserves all data and metadata)
            import shutil
            filename = os.path.basename(filepath)
            deleted_path = os.path.join(deleted_dir, filename)
            
            # Handle name collision
            counter = 1
            while os.path.exists(deleted_path):
                base, ext = os.path.splitext(filename)
                deleted_path = os.path.join(deleted_dir, f"{base}_{counter}{ext}")
                counter += 1
            
            shutil.move(filepath, deleted_path)
            
            # Return the deleted path as "backup" for potential recovery
            backup_path = deleted_path
        
        # Production deletion (immediate, unrecoverable)
        else:
            os.remove(filepath)
        
        return True, "", backup_path
    
    except Exception as e:
        return False, str(e), None


